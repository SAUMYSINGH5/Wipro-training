from openpyxl import Workbook, load_workbook
import os

def write_excel(filename):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Name", "Age"])
    sheet.append(["John Doe", 30])
    sheet.append(["Jane Smith", 25])
    workbook.save(filename)

def read_excel(filename):
    workbook = load_workbook(filename)
    sheet = workbook.active
    for row in sheet.iter_rows(values_only=True):
        print(f"Name: {row[0]}, Age: {row[1]}")

def delete_excel(filename):
    if os.path.exists(filename):
        os.remove(filename)
        print(f"{filename} deleted Successfully")
    else:
        print(f"{filename} does not exist")

filename = "myfile.xlsx"
write_excel(filename)
print('Data read from EXCEL file: ')
read_excel(filename)
delete_excel(filename)

















'''from openpyxl import Workbook, load_workbook
import os

def write_excel(filename):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Name", "Age"])
    sheet.append(["John Doe", 30])
    sheet.append(["Jame Smith", 25])
    sheet.save(filename)

def read_excel(filename):
     workbook = load_workbook(filename)
     sheet = workbook.active
     for row in sheet.iter_rows(values_only=True):
         print(f"Name: {row[0]}, Age: {row[1]}")

def delete_excel(filename):
    if os.path.exists(filename):
        os.remove(filename)
        print(f"{filename} deleted Sucessfully")
    else:
        print(f"{filename} does not exist")

filename = "myfile.excel"
write_excel(filename)
print('Data read from EXCEL file: ')
read_excel(filename)
delete_excel(filename)'''